from typing import Dict, Any
from .base import BaseDocumentHandler

class ExcelHandler(BaseDocumentHandler):
    """Handler for Excel spreadsheets (.xlsx) using openpyxl."""

    def __init__(self, file_path: str):
        super().__init__(file_path)
        self._wb = None
        self._text = ""
        self._metadata = {}
        self._sheet_data = {}  # sheet_name -> list of rows

    def load(self) -> None:
        """Load and parse the XLSX file."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is not installed. Please install it via 'pip install openpyxl' to use ExcelHandler.")

        try:
            self._wb = openpyxl.load_workbook(self.file_path, data_only=True)
            
            text_parts = []
            self._sheet_data = {}
            
            for sheet_name in self._wb.sheetnames:
                sheet = self._wb[sheet_name]
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):
                        rows.append(row_data)
                        text_parts.append("\t".join(row_data))
                
                self._sheet_data[sheet_name] = rows
                text_parts.append("")

            self._text = "\n".join(text_parts)
            
            props = self._wb.properties
            self._metadata = {
                "author": props.creator,
                "created": str(props.created),
                "modified": str(props.modified),
                "title": props.title,
                "subject": props.subject,
                "keywords": props.keywords,
                "category": props.category,
                "sheet_names": self._wb.sheetnames
            }
            self._metadata = {k: v for k, v in self._metadata.items() if v is not None}

        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file {self.file_path}: {e}")

    def get_text(self) -> str:
        if self._wb is None:
            self.load()
        return self._text

    def get_metadata(self) -> Dict[str, Any]:
        if self._wb is None:
            self.load()
        return self._metadata

    def to_markdown(self) -> str:
        """Convert Excel to Markdown with proper table formatting per sheet."""
        if self._wb is None:
            self.load()
        
        md_parts = [f"# Document: {__import__('os').path.basename(self.file_path)}\n"]
        
        if self._metadata:
            md_parts.append("## Metadata\n")
            for key, value in self._metadata.items():
                md_parts.append(f"- **{key}**: {value}")
            md_parts.append("")
        
        for sheet_name, rows in self._sheet_data.items():
            if not rows:
                continue
            
            md_parts.append(f"## Sheet: {sheet_name}\n")
            
            header = rows[0]
            col_count = len(header)
            
            md_table = "| " + " | ".join(header) + " |\n"
            md_table += "| " + " | ".join(["---"] * col_count) + " |\n"
            
            for row in rows[1:]:
                padded = row + [""] * (col_count - len(row))
                final = padded[:col_count]
                md_table += "| " + " | ".join(final) + " |\n"
            
            md_parts.append(md_table)
            md_parts.append("")
        
        return "\n".join(md_parts)
