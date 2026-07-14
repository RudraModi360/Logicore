"""
Excel Operations Helper Scripts

Tool functions for the excel_operations skill.
Provides CRUD operations on Excel workbooks and sheets.
"""

from pathlib import Path
from typing import Optional, Dict, Any, List

_EXCEL_EXTENSIONS = {'.xlsx', '.xlsm', '.xltx', '.xltm'}


def _validate_excel_path(file_path: str) -> Optional[Dict[str, Any]]:
    """Validate path points to an existing Excel file. Returns error dict or None."""
    path = Path(file_path)
    if not path.exists():
        return {"error": f"File not found: {file_path}"}
    if path.is_dir():
        return {"error": f"Expected a file, got a directory: {file_path}. Use list_files to explore directories."}
    if path.suffix.lower() not in _EXCEL_EXTENSIONS:
        return {"error": f"Not an Excel file: {file_path}. Expected one of: {', '.join(sorted(_EXCEL_EXTENSIONS))}"}
    return None


def create_workbook(file_path: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
    """
    Create a new Excel workbook with an initial sheet.

    Args:
        file_path: Path where the .xlsx file will be saved
        sheet_name: Name of the first sheet (default: Sheet1)

    Returns:
        Dict with status and file info
    """
    from openpyxl import Workbook

    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(str(path))

    return {
        "status": "created",
        "file_path": str(path.resolve()),
        "sheet_name": sheet_name,
        "size_bytes": path.stat().st_size,
    }


def read_sheet(file_path: str, sheet_name: Optional[str] = None, max_rows: int = 100) -> Dict[str, Any]:
    """
    Read data from an Excel sheet and return as rows.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of sheet to read (default: active sheet)
        max_rows: Maximum rows to read (default: 100)

    Returns:
        Dict with headers and data rows
    """
    from openpyxl import load_workbook

    err = _validate_excel_path(file_path)
    if err:
        return err

    path = Path(file_path)
    wb = load_workbook(str(path), read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
        else:
            if i > max_rows:
                break
            rows.append([str(c) if c is not None else "" for c in row])

    wb.close()

    return {
        "sheet_name": sheet_name,
        "headers": headers,
        "rows": rows,
        "total_rows_read": len(rows),
        "truncated": ws.max_row is not None and len(rows) >= max_rows,
    }


def list_sheets(file_path: str) -> Dict[str, Any]:
    """
    List all sheet names in an Excel workbook.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        Dict with list of sheet names
    """
    from openpyxl import load_workbook

    err = _validate_excel_path(file_path)
    if err:
        return err

    path = Path(file_path)

    wb = load_workbook(str(path), read_only=True)
    sheets = wb.sheetnames
    wb.close()

    return {"file_path": str(path.resolve()), "sheet_names": sheets, "sheet_count": len(sheets)}


def write_cells(file_path: str, sheet_name: str, start_row: int, start_col: int, data: List[List[Any]], overwrite: bool = True) -> Dict[str, Any]:
    """
    Write a block of data to specific cells in a sheet.

    Args:
        file_path: Path to the .xlsx file (must exist)
        sheet_name: Name of the target sheet
        start_row: Starting row number (1-indexed)
        start_col: Starting column number (1-indexed, A=1)
        data: 2D list of values to write
        overwrite: Whether to overwrite existing data (default: True)

    Returns:
        Dict with status and cells written
    """
    from openpyxl import load_workbook

    err = _validate_excel_path(file_path)
    if err:
        return err

    path = Path(file_path)
    wb = load_workbook(str(path))
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}

    ws = wb[sheet_name]
    cells_written = 0

    for r_idx, row in enumerate(data):
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx)
            if not overwrite and cell.value is not None:
                continue
            cell.value = value
            cells_written += 1

    wb.save(str(path))
    wb.close()

    return {
        "status": "written",
        "sheet_name": sheet_name,
        "cells_written": cells_written,
        "range": f"{_col_letter(start_col)}{start_row}:{_col_letter(start_col + len(data[0]) - 1)}{start_row + len(data) - 1}" if data else "empty",
    }


def format_cells(file_path: str, sheet_name: str, cell_range: str, bold: Optional[bool] = None, font_size: Optional[int] = None, font_color: Optional[str] = None, bg_color: Optional[str] = None) -> Dict[str, Any]:
    """
    Apply formatting to a range of cells.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name of the target sheet
        cell_range: Cell range string (e.g., "A1:D1")
        bold: Set bold formatting (True/False)
        font_size: Set font size
        font_color: Set font color (hex like "FF0000")
        bg_color: Set background color (hex like "FFFF00")

    Returns:
        Dict with status and formatted range
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    err = _validate_excel_path(file_path)
    if err:
        return err

    path = Path(file_path)
    wb = load_workbook(str(path))
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"error": f"Sheet '{sheet_name}' not found."}

    ws = wb[sheet_name]
    cells_formatted = 0

    for row in ws[cell_range]:
        for cell in row:
            if bold is not None:
                cell.font = Font(bold=bold, size=cell.font.size if not font_size else font_size,
                                 color=cell.font.color if not font_color else font_color)
            if font_size is not None:
                cell.font = Font(size=font_size, bold=cell.font.bold,
                                 color=cell.font.color if not font_color else font_color)
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cells_formatted += 1

    wb.save(str(path))
    wb.close()

    return {"status": "formatted", "sheet_name": sheet_name, "range": cell_range, "cells_formatted": cells_formatted}


def add_sheet(file_path: str, sheet_name: str) -> Dict[str, Any]:
    """
    Add a new sheet to an existing workbook.

    Args:
        file_path: Path to the .xlsx file
        sheet_name: Name for the new sheet

    Returns:
        Dict with status
    """
    from openpyxl import load_workbook

    err = _validate_excel_path(file_path)
    if err:
        return err

    path = Path(file_path)
    wb = load_workbook(str(path))
    if sheet_name in wb.sheetnames:
        wb.close()
        return {"error": f"Sheet '{sheet_name}' already exists."}

    wb.create_sheet(sheet_name)
    wb.save(str(path))
    wb.close()

    return {"status": "added", "sheet_name": sheet_name, "all_sheets": wb.sheetnames}


def get_summary(file_path: str) -> Dict[str, Any]:
    """
    Get a summary of the workbook structure.

    Args:
        file_path: Path to the .xlsx file

    Returns:
        Dict with sheet count, names, and dimensions
    """
    from openpyxl import load_workbook

    err = _validate_excel_path(file_path)
    if err:
        return err

    path = Path(file_path)
    wb = load_workbook(str(path), read_only=True)
    sheets_info = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets_info.append({
            "name": name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })
    wb.close()

    return {
        "file_path": str(path.resolve()),
        "sheet_count": len(sheets_info),
        "sheets": sheets_info,
        "size_bytes": path.stat().st_size,
    }


def _col_letter(col_num: int) -> str:
    """Convert 1-indexed column number to Excel letter (1=A, 26=Z, 27=AA)."""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + col_num % 26) + result
        col_num //= 26
    return result
