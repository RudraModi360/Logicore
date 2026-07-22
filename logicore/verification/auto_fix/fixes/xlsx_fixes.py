"""
XLSX auto-fixes.

Handles fixing common Excel issues:
- Add headers to headerless sheets
- Remove empty sheets
- Fix CSV formatting
"""

from __future__ import annotations

import logging
import os
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from logicore.verification.auto_fix import AutoFixEngine

logger = logging.getLogger(__name__)


def register_fixes(engine: "AutoFixEngine") -> None:
    """Register XLSX fix handlers with the auto-fix engine."""
    engine.register_fix("content", "no header row", _fix_add_headers)
    engine.register_fix("content", "empty", _fix_remove_empty_sheets)
    engine.register_fix("format", "inconsistent column", _fix_csv_consistency)


def _fix_add_headers(artifact_path: str, issue) -> bool:
    """Add generic headers to the first row if missing."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    try:
        ext = os.path.splitext(artifact_path)[1].lower()
        if ext == ".csv":
            return _fix_csv_headers(artifact_path)

        wb = load_workbook(artifact_path)
        modified = False

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Check if first row has headers.
            if sheet.max_row and sheet.max_row >= 1:
                first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if first_row:
                    headers = first_row[0]
                    # Check if headers are all None/empty.
                    if all(h is None or str(h).strip() == "" for h in headers):
                        # Add generic headers.
                        for col_idx in range(len(headers)):
                            cell = sheet.cell(row=1, column=col_idx + 1)
                            cell.value = f"Column {col_idx + 1}"
                        modified = True

        if modified:
            wb.save(artifact_path)

        return modified

    except Exception as exc:
        logger.debug(f"Failed to add headers: {exc}")
        return False


def _fix_csv_headers(artifact_path: str) -> bool:
    """Add headers to CSV if first row is empty."""
    try:
        with open(artifact_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        if not lines:
            return False

        # Check if first row is empty.
        first_row = lines[0].strip()
        if first_row:
            return False  # Already has headers.

        # Count columns from second row.
        if len(lines) > 1:
            import csv
            reader = csv.reader([lines[1]])
            row = next(reader)
            num_cols = len(row)
        else:
            num_cols = 3  # Default.

        # Add generic headers.
        headers = ",".join([f"Column {i+1}" for i in range(num_cols)])
        lines[0] = headers + "\n"

        with open(artifact_path, "w", encoding="utf-8") as f:
            f.writelines(lines)

        return True

    except Exception as exc:
        logger.debug(f"Failed to fix CSV headers: {exc}")
        return False


def _fix_remove_empty_sheets(artifact_path: str, issue) -> bool:
    """Remove empty sheets from workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    try:
        wb = load_workbook(artifact_path)

        if len(wb.sheetnames) <= 1:
            return False  # Don't remove if only one sheet.

        sheets_to_remove = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Check if sheet has any data.
            if sheet.max_row is None or sheet.max_row == 0:
                sheets_to_remove.append(sheet_name)
            elif sheet.max_row == 1:
                # Check if only header row.
                first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if first_row and all(h is None for h in first_row[0]):
                    sheets_to_remove.append(sheet_name)

        if not sheets_to_remove:
            return False

        # Remove sheets.
        for sheet_name in sheets_to_remove:
            del wb[sheet_name]

        wb.save(artifact_path)
        return True

    except Exception as exc:
        logger.debug(f"Failed to remove empty sheets: {exc}")
        return False


def _fix_csv_consistency(artifact_path: str, issue) -> bool:
    """Fix inconsistent column counts in CSV by padding shorter rows."""
    try:
        import csv
    except ImportError:
        return False

    try:
        with open(artifact_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            return False

        # Find max column count.
        max_cols = max(len(row) for row in rows)

        # Pad shorter rows.
        modified = False
        for i, row in enumerate(rows):
            if len(row) < max_cols:
                rows[i] = row + [""] * (max_cols - len(row))
                modified = True

        if not modified:
            return False

        # Write back.
        with open(artifact_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)

        return True

    except Exception as exc:
        logger.debug(f"Failed to fix CSV consistency: {exc}")
        return False
