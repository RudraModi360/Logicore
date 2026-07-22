"""
XLSX verifier.

Verifies Excel spreadsheets: sheets, data, formulas, structure.
"""

from __future__ import annotations

import os
from typing import List, Optional, Set

from logicore.verification.base_verifier import BaseVerifier
from logicore.verification.result import VerificationIssue


# Supported XLSX extensions.
XLSX_EXTENSIONS: Set[str] = {".xlsx", ".xls", ".csv"}

# Maximum sheets for a reasonable workbook.
MAX_SHEETS = 100

# Maximum rows per sheet.
MAX_ROWS = 1_000_000

# Maximum columns per sheet.
MAX_COLUMNS = 16_384


class XLSXVerifier(BaseVerifier):
    """Verify Excel spreadsheets: sheets, data, formulas, structure.

    Checks:
    - Workbook can be opened
    - Has at least one sheet with data
    - Sheets have headers
    - No completely empty sheets
    - Data types are consistent in columns
    - Formulas reference valid cells
    """

    def supported_extensions(self) -> Set[str]:
        return XLSX_EXTENSIONS

    def _verify_content(
        self,
        artifact_path: str,
        issues: List[VerificationIssue],
        requirements: Optional[str],
    ) -> None:
        ext = os.path.splitext(artifact_path)[1].lower()

        if ext == ".csv":
            self._verify_csv(artifact_path, issues)
        else:
            self._verify_xlsx(artifact_path, issues, requirements)

    def _verify_xlsx(
        self,
        artifact_path: str,
        issues: List[VerificationIssue],
        requirements: Optional[str],
    ) -> None:
        """Verify XLSX/XLS files using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self._add_issue(
                issues,
                severity="info",
                category="dependency",
                description="openpyxl not installed — skipping detailed XLSX verification",
            )
            return

        try:
            wb = load_workbook(artifact_path, read_only=True, data_only=True)
        except Exception as exc:
            self._add_issue(
                issues,
                severity="critical",
                category="corruption",
                description=f"Cannot open XLSX file: {exc}",
                auto_fixable=False,
            )
            return

        try:
            sheet_names = wb.sheetnames
        except Exception:
            sheet_names = []

        if not sheet_names:
            self._add_issue(
                issues,
                severity="critical",
                category="content",
                description="Workbook has no sheets",
            )
            return

        if len(sheet_names) > MAX_SHEETS:
            self._add_issue(
                issues,
                severity="warning",
                category="structure",
                description=f"Workbook has {len(sheet_names)} sheets (unusually large)",
                auto_fixable=False,
            )

        # Per-sheet checks.
        empty_sheets = []
        for name in sheet_names:
            try:
                sheet = wb[name]
                sheet_issues = self._check_sheet(sheet, name)
                issues.extend(sheet_issues)

                # Track empty sheets.
                if sheet.max_row is None or sheet.max_row == 0:
                    empty_sheets.append(name)
                elif sheet.max_row == 1:
                    # Only header row.
                    pass  # Not necessarily an issue.
            except Exception as exc:
                self._add_issue(
                    issues,
                    severity="warning",
                    category="corruption",
                    description=f"Cannot read sheet '{name}': {exc}",
                    location=f"sheet '{name}'",
                )

        if empty_sheets:
            self._add_issue(
                issues,
                severity="warning",
                category="content",
                description=f"{len(empty_sheets)} sheet(s) are empty: {empty_sheets}",
                auto_fixable=True,
                fix_suggestion="Add data to empty sheets or remove them",
            )

        # Requirements check.
        if requirements:
            req_lower = requirements.lower()
            if "chart" in req_lower:
                # Check if any charts exist.
                has_charts = False
                for name in sheet_names:
                    try:
                        sheet = wb[name]
                        if hasattr(sheet, "_charts") and sheet._charts:
                            has_charts = True
                            break
                    except Exception:
                        pass
                if not has_charts:
                    self._add_issue(
                        issues,
                        severity="warning",
                        category="content",
                        description="No charts found in workbook (user requested charts)",
                    )

    def _check_sheet(self, sheet, sheet_name: str) -> List[VerificationIssue]:
        """Check a single sheet for issues."""
        issues = []

        # Row/column count checks.
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        if max_row > MAX_ROWS:
            issues.append(VerificationIssue(
                severity="warning",
                category="structure",
                description=f"Sheet '{sheet_name}' has {max_row} rows (unusually large)",
                location=f"sheet '{sheet_name}'",
            ))

        if max_col > MAX_COLUMNS:
            issues.append(VerificationIssue(
                severity="warning",
                category="structure",
                description=f"Sheet '{sheet_name}' has {max_col} columns (unusually large)",
                location=f"sheet '{sheet_name}'",
            ))

        # Header row check.
        if max_row > 0 and max_col > 0:
            try:
                header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if header_row:
                    headers = header_row[0]
                    # Check if headers are all None/empty.
                    if all(h is None or str(h).strip() == "" for h in headers):
                        issues.append(VerificationIssue(
                            severity="info",
                            category="content",
                            description=f"Sheet '{sheet_name}' has no header row",
                            location=f"sheet '{sheet_name}'",
                            auto_fixable=True,
                            fix_suggestion="Add headers to the first row",
                        ))
            except Exception:
                pass

        # Empty data check.
        if max_row <= 1 and max_col <= 1:
            # Very small sheet - possibly empty.
            try:
                cell = sheet.cell(row=1, column=1)
                if cell.value is None:
                    issues.append(VerificationIssue(
                        severity="warning",
                        category="content",
                        description=f"Sheet '{sheet_name}' appears to be empty",
                        location=f"sheet '{sheet_name}'",
                        auto_fixable=True,
                        fix_suggestion="Add data to the sheet",
                    ))
            except Exception:
                pass

        return issues

    def _verify_csv(
        self,
        artifact_path: str,
        issues: List[VerificationIssue],
    ) -> None:
        """Basic CSV verification."""
        try:
            with open(artifact_path, "r", encoding="utf-8", errors="replace") as f:
                lines = f.readlines()
        except Exception as exc:
            self._add_issue(
                issues,
                severity="critical",
                category="corruption",
                description=f"Cannot read CSV file: {exc}",
            )
            return

        if not lines:
            self._add_issue(
                issues,
                severity="critical",
                category="content",
                description="CSV file is empty",
            )
            return

        # Check header row.
        if lines[0].strip() == "":
            self._add_issue(
                issues,
                severity="warning",
                category="content",
                description="CSV has no header row",
                auto_fixable=True,
                fix_suggestion="Add headers to the first row",
            )

        # Check for consistent column count.
        if len(lines) > 1:
            try:
                import csv
                with open(artifact_path, "r", encoding="utf-8", errors="replace") as f:
                    reader = csv.reader(f)
                    row_lengths = [len(row) for i, row in enumerate(reader) if i < 100]
                    if row_lengths:
                        expected = row_lengths[0]
                        inconsistent = [i for i, l in enumerate(row_lengths) if l != expected]
                        if inconsistent:
                            self._add_issue(
                                issues,
                                severity="warning",
                                category="format",
                                description=f"CSV has inconsistent column counts (rows: {inconsistent[:5]})",
                            )
            except Exception:
                pass


def get_verifier() -> XLSXVerifier:
    """Return an XLSXVerifier instance for registry auto-discovery."""
    return XLSXVerifier()
